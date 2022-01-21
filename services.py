
from tkinter import *
from tkinter import Entry, Label, Tk, ttk, messagebox
import tkinter as tk
from openpyxl import load_workbook
from datetime import date, datetime
from tkcalendar import DateEntry
from babel.numbers import *

file = 'العملاء.xlsx' 

wb = load_workbook(file)

wb.security.workbookPassword = '...'
wb.security.lockStructure = True

services = ['اتعاب لجنه داخلية','اضافة سياره','اعداد و مراجعة ميزانيه','اقرار ضرائب عامه'
        ,'اقرار ضرائب قيمه مضافه', 'اقرار ضرائب مرتبات', 'بطاقه ضريبيه'
        ,'تجديد اشتراك البوابه الالكترونيه', 'تحت الحساب','تسوية ملف ضريبى', 'تعديل النشاط',
        'جواب مرور', 'حفظ الملف بالضرائب','رسوم استخراج مستوردين', 'رسوم البوابه الالكترونيه'
        ,'رسوم تجديد مستوردين', 'سجل تجارى','سداد ضرائب عامه'
        , 'سداد ضرائب قيمه مضافه','شطب سجل تجارى', 'شهاده بالموقف الضريبى'
        ,'شهادة دخل 1', 'شهادة دخل 2', 'عمل موقع الكترونى', 'غرفه تجاريه',
        'فحص ضرائب عامه', 'فحص ضريبة قيمه مضافه','لجنة طعن ضرائب عامه', 'مركز مالى',
        'ميزانيه عموميه', 'نماذج 41 ض',]

ws = wb.create_sheet('الخدمات')
for i in services:
    ws.append([i])
wb.save(file)    
