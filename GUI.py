from tkinter import *
from tkinter import Entry, Label, Tk, ttk, messagebox
import tkinter as tk
from openpyxl import load_workbook
from datetime import date, datetime
from tkcalendar import DateEntry
from babel.numbers import *

# Get Resolution of Screen.
root = tk.Tk()
width = int(root.winfo_screenwidth() * 0.9)
height = int(root.winfo_screenheight() * 0.9)
root.destroy()

# Load Excel Files.
file = 'العملاء.xlsx' 
file0 = 'المصروفات.xlsx'
try:
        wb = load_workbook(file)
        wb0 = load_workbook(file0)
except:
        messagebox.showerror('Files Not Found!',' ملفات الاكسيل غير موجوده')
        print('Files do not exist')


# services = ['اتعاب لجنه داخلية','اضافة سياره','اعداد و مراجعة ميزانيه','اقرار ضرائب عامه'
#         ,'اقرار ضرائب قيمه مضافه', 'اقرار ضرائب مرتبات', 'بطاقه ضريبيه'
#         ,'تجديد اشتراك البوابه الالكترونيه', 'تحت الحساب','تسوية ملف ضريبى', 'تعديل النشاط',
#         'جواب مرور', 'حفظ الملف بالضرائب','رسوم استخراج مستوردين', 'رسوم البوابه الالكترونيه'
#         ,'رسوم تجديد مستوردين', 'سجل تجارى','سداد ضرائب عامه'
#         , 'سداد ضرائب قيمه مضافه','شطب سجل تجارى', 'شهاده بالموقف الضريبى'
#         ,'شهادة دخل 1', 'شهادة دخل 2', 'عمل موقع الكترونى', 'غرفه تجاريه',
#         'فحص ضرائب عامه', 'فحص ضريبة قيمه مضافه','لجنة طعن ضرائب عامه', 'مركز مالى',
#         'ميزانيه عموميه', 'نماذج 41 ض',]


# Fonts
font22 = 'fantasy 22 bold'
font20 = 'fantasy 20 bold'
font18 = 'fantasy 18 bold'
font10 = 'fantasy 10 bold'

#Colors
baseColor = '#E5E8C7'
labelColor = 'black'

class Edit_Client(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)

        self.geometry(f'{width}x{height}')
        self.title('تعديل بيانات عميل')
        self['bg']=baseColor


        entries = []
        for sheet in wb.worksheets:
                if sheet.cell(row=1,column=1).value != None and sheet.title != 'الخدمات':
                        entries.append(f'{str(sheet.cell(row=1,column=1).value)} ({str(sheet.cell(row=1,column=2).value)})')
                        
        Label(self, text='اختر العميل',
        bg = baseColor, fg = labelColor ,font = font20 ).place(relx=.5, rely=.05,anchor= CENTER)

        codeentries = []
        for sheet in wb.worksheets:
                if sheet.cell(row=1,column=2).value != None :
                        codeentries.append(sheet.cell(row=1,column=2).value)                

        def select_input_names(event):
                value = event.widget.get()
                if value == '':
                        menu['values'] = entries
                else:
                        data = []
                        for item in entries:
                                if value.lower() in item.lower():
                                        data.append(item)

                        menu['values'] = data

        name = StringVar()
        menu = ttk.Combobox(self, textvariable= name, font=font20,values = entries )
        app.option_add('*TCombobox*Listbox.font', font20)
        menu.place(relx=.5, rely=.1,anchor= CENTER, width=int(width*0.417))
        menu['values'] = entries
        menu.bind('<KeyRelease>', select_input_names)

        Label(self, text='.برجاء عدم ترك الخانه فارغه',
        bg = baseColor, fg = labelColor ,font = font10).place(relx=.65, rely=.15,anchor= CENTER)


        Label(self, text='قم بادخال البيانات المراد تعديلها',
        bg = baseColor, fg = '#E85E3F' ,font = 'fantasy 25 bold').place(relx=.5, rely=.25,anchor= CENTER,)


        Label(self, text='الاسم',
        bg = baseColor, fg = labelColor ,font = font20 ).place(relx=.9, rely=.35,anchor= CENTER, width=int(width*0.138))

        edit_name = tk.StringVar()
        entry1 = ttk.Entry(self, textvariable=edit_name, justify = LEFT, font = font20)
        entry1.place(relx=.65, rely=.35,anchor= CENTER, width=int(width*0.4))


        Label(self, text='الكود',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.9, rely=.65,anchor= CENTER, width=int(width*0.138))

        code = IntVar()
        entry1 = ttk.Entry(self, textvariable=code, justify = LEFT, font = font20)
        entry1.place(relx=.77, rely=.65,anchor= CENTER, width=int(width*0.138))


        Label(self, text='رقم التسجيل',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.9, rely=.75,anchor= CENTER, width=int(width*0.138))

        record = tk.IntVar()
        entry1 = ttk.Entry(self, textvariable=record, justify = LEFT, font = font20)
        entry1.place(relx=.77, rely=.75,anchor= CENTER, width=int(width*0.138))


        Label(self, text='رقم التليفون',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.9, rely=.45,anchor= CENTER, width=int(width*0.1))

        phone = tk.StringVar()
        entry1 = ttk.Entry(self, textvariable=phone, justify = LEFT, font = font20)
        entry1.place(relx=.65, rely=.45,anchor= CENTER, width=int(width*0.4))


        Label(self, text='العنوان',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.9, rely=.55,anchor= CENTER, width=int(width*0.1))

        address  = tk.StringVar()
        entry1 = ttk.Entry(self, textvariable=address, justify = LEFT, font = font20)
        entry1.place(relx=.65, rely=.55,anchor= CENTER, width=int(width*0.4))


        def edit():
                try:
                        ws = wb[name.get()]
                        ws0 = wb0[name.get()]
                except:
                        messagebox.showerror('Not Exists!','الاسم غير موجود')
                        print('cant find worksheet')
                        return False 
                try:
                        name.get() and code.get() and record.get() and phone.get() and address.get() 
                except:
                        messagebox.showerror('Invalid!','من فضلك قم بادخال الخانات بطريقه صحيحه ')
                        print('invalid input')
                        return False

                # not None Validation
                if name.get() == None or code.get() == None :
                        messagebox.showerror('Invalid!','من فضلك قم بادخال الخانات المطلوبه ') 
                        print('None input')
                        return False

                # not "" validation        
                if name.get() == "" :
                        messagebox.showerror('Invalid!','من فضلك نأكد من ادخال الاسم  ')
                        print('No Name')
                        return False
                
                # phone validation
                if not phone.get().isnumeric:
                        messagebox.showerror('Invalid!','من فضلك نأكد من ادخال التليفون بطريقه صحيحه ')
                        print('Phone is not numeric')
                        return False

                # Case validations

                if code.get() in codeentries:
                        # self.destroy()
                        messagebox.showerror('Already Exists!','الكود موجود بالفعل ') 
                        print('code already exists')
                        return False

                else:
                        if edit_name.get() != '':
                                ws['A1'] = edit_name.get()
                                ws.title = f'{edit_name.get()} ({ws["B1"].value})'

                        if code.get() != 0: 
                                ws['B1'] = code.get()
                                ws.title = f'{ws["A1"].value} ({code.get()})'

                        if edit_name.get() != '' and code != 0:
                                ws.title = f'{edit_name.get()} ({code.get()})'        

                        if record.get() != 0:        
                                ws['C1'] = record.get()

                        if phone.get() != '':
                                ws['D1'] = phone.get()

                        if address.get() != '':
                                ws['E1'] = address.get()
                        wb.save(file)

                        if edit_name.get() != '':
                                ws0['A1'] = edit_name.get()
                                ws0.title = f'{edit_name.get()} ({ws0["B1"].value})'

                        if code.get() != 0: 
                                ws0['B1'] = code.get()
                                ws0.title = f'{ws0["A1"].value} ({code.get()})'

                        if edit_name.get() != '' and code != 0:
                                ws0.title = f'{edit_name.get()} ({code.get()})'

                        if record.get() != 0:        
                                ws0['C1'] = record.get()

                        if phone.get() != '':
                                ws0['D1'] = phone.get()

                        if address.get() != '':
                                ws0['E1'] = address.get()
                        wb0.save(file0)
                        messagebox.showinfo('Done','تم التعديل بنجاح ')


        Button(self, height = 1, width = 10, bg = '#05659E', fg = 'white',
        activebackground='#43516C', font = font20, bd = '4px solid #DBA531', 
                text='تعديل',
                command=edit).place(relx=.5, rely=.9,anchor= CENTER)
        
        
        Button(self, height = 1, width = 8, bg = 'grey', fg = 'white',
        activebackground='#43516C', font = 'fantasy 15 bold', bd = '4px solid #DBA531', 
                text='اغلاق',
                command=self.destroy).place(relx=.1, rely=.95,anchor= CENTER)


class New_Service(tk.Toplevel):
      def __init__(self, parent):
        super().__init__(parent)

        self.geometry(f'{width}x{height}')
        self.title('اضافة خدمات جديده')
        self['bg']=baseColor

        
        Label(self, text='قائمة الخدمات',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.5, rely=.3,anchor= CENTER, width=int(width*0.14))

        
        def select_input_services(event):
                value = event.widget.get()
                if value == '':
                        menu2['values'] = services
                else:
                        data = []
                        for item in services:
                                if value.lower() in item.lower():
                                        data.append(item)

                        menu2['values'] = data
        
        services = []
        ws = wb['الخدمات']
        for cell in ws['A']:
                if cell.value != None or cell.value != '':
                        services.append(cell.value)

        new_service = tk.StringVar()
        menu2 = ttk.Combobox(self, textvariable = new_service, font=font20, values = services) 
        app.option_add('*TCombobox*Listbox.font', font20)
        menu2.place(relx=.5, rely=.4,anchor= CENTER, width=int(width*0.5))
        menu2.bind('<KeyRelease>', select_input_services)
        

        def newService():
                try:
                        new_service.get()
                
                        if new_service.get() == '':
                                messagebox.showerror('Invalid!','من فضلك قم بادخال الخانات  المطلوبه ')
                                print('No Service')
                                return False

                        if new_service.get() not in services:
                                ws = wb['الخدمات']
                                ws.append([new_service.get()])
                                wb.save(file)
                                self.destroy()
                                messagebox.showinfo('Done!','تم اضافة خدمه جديده')

                        if new_service.get() in services:
                                messagebox.showerror('already exists!','هذه الخدمه موجوده بالفعل')
                except:
                        messagebox.showerror('Invalid!','من فضلك قم بادخال الخانات بطريقه صحيحه ') 
                        print('invalid service')
                        return False


        def deleteService():
                try:
                        new_service.get()
                
                        if new_service.get() == '':
                                messagebox.showerror('Invalid!','من فضلك قم بادخال الخانات  المطلوبه ')
                                print('No Service')
                                return False

                        if new_service.get() in services:
                                ws = wb['الخدمات']
                                for i in range(1,ws.max_row+1):
                                        if cell(column=1, row=i).value == new_service.get():
                                                ws.delete_rows(i,1)
                                                wb.save(file)     
                                                self.destroy()            
                                                messagebox.showinfo('Done!','تم الحذف')
                        if new_service.get() not in services:
                                messagebox.showerror('not exists!','هذه الخدمه غير موجوده ')
                                print('Service already exists')
                                return False
                except:
                        messagebox.showerror('Invalid!','من فضلك قم بادخال الخانات بطريقه صحيحه ') 
                        print('invalid ervice')
                        return False



        Button(self, height = 1, width = 12, bg = 'green', fg = 'white',
        activebackground='#43516C', font = font20, bd = '8px solid #DBA531', 
                text='اضافه',
                command=newService).place(relx=.4, rely=.55,anchor= CENTER)

        Button(self, height = 1, width = 12, bg = 'red', fg = 'white',
        activebackground='#43516C', font = font20, bd = '8px solid #DBA531', 
                text='حذف',
                command=deleteService).place(relx=.6, rely=.55,anchor= CENTER)        
        
        Button(self, height = 1, width = 10, bg = 'grey', fg = 'white',
        activebackground='#43516C', font = 'fantasy 15 bold', bd = '8px solid #DBA531', 
                text='اغلاق',
                command=self.destroy).place(relx=.5, rely=.8,anchor= CENTER)

        

class Expenses_View(tk.Toplevel):
      def __init__(self, parent):
        super().__init__(parent)

        self.geometry(f'{width}x{height}')
        self.title('المصروفات')
        self['bg']=baseColor

        Label(self, text='من',
        bg = '#43516C', fg = 'white', font = font20,
        borderwidth=5, relief="ridge", padx=20, pady=10).place(relx=.7, rely=.4,anchor= CENTER, width=int(width*0.28))

        startdate = DateEntry(self,width=30,bg="darkblue",fg="white", font = font18,
        year=datetime.now().year,month=datetime.now().month,day=datetime.now().day
        ,locale='en_US', date_pattern='dd/MM/yyyy')
        startdate.place(relx=.7, rely=.5,anchor= CENTER, width=int(width*0.14), height=50)
        

        Label(self, text='الى',
        bg = '#43516C', fg = 'white', font = font20,
        borderwidth=5, relief="ridge", padx=20, pady=10).place(relx=.3, rely=.4,anchor= CENTER, width=int(width*0.28))

        enddate = DateEntry(self,width=30,bg="darkblue",fg="white", font = font18,
        year=datetime.now().year,month=datetime.now().month,day=datetime.now().day
        ,locale='en_US', date_pattern='dd/MM/yyyy')
        enddate.place(relx=.3, rely=.5,anchor= CENTER, width=int(width*0.14), height=50)
        
        
        def viewExpenses():
                treev = ttk.Treeview(self, selectmode ='browse', style="mystyle.Treeview", height=height)
                treev.pack()
                style = ttk.Style()
                style.configure("mystyle.Treeview", background = baseColor ,rowheight=100,
                highlightthickness=0, bd=0, font=('Helvetica', 14)) 
                style.configure("mystyle.Treeview.Heading", font=font20) 
                style.layout("mystyle.Treeview", [('mystyle.Treeview.treearea', {'sticky': 'nswe'})]) 
                verscrlbar = ttk.Scrollbar(self,
                                orient ="vertical",
                                command = treev.yview)
                verscrlbar.pack(side ='left', fill ='x')   
                treev.configure(xscrollcommand = verscrlbar.set)
                treev["columns"] = ("1", "2")
                treev['show'] = 'headings'
                treev.column("1", width = int(width*0.3),anchor ='e')
                treev.column("2", width = int(width*0.7),anchor ='e')
                treev.heading("1", text ="المصروف")
                treev.heading("2", text ="الحساب")


                expenses = []
                try:
                        ws0 = wb0['مصروفات اداريه']
                except:
                        messagebox.showinfo('Not Found!','لا يوجد مصروفات اداريه حتى الان')
                        print('No Manegerial Expenses')  
                        return False

                expense = ''
                c = 0
                for i in range (1, ws0.max_row+1):
                        # Check if row date exists.
                        try:
                                celldate = date(
                                ws0.cell(row=i, column=4).value,
                                ws0.cell(row=i, column=5).value,
                                ws0.cell(row=i, column=6).value)
                        except:
                                continue
                        # New Expense.
                        if ws0.cell(row=i,column=3).value not in expenses :
                                c += 1
                                expenses.append(ws0.cell(row=i,column=3).value)
                                expense = ws0.cell(row=i,column=3).value
                                amount = 0 
                                for i in range (1, ws0.max_row+1):
                                        if ws0.cell(row=i,column=3).value == expense:
                                                # If Date Matches.
                                                if celldate >= startdate.get_date() and celldate <= enddate.get_date():
                                                        amount += int(ws0.cell(row=i,column=2).value)
                                if c%2 == 0:                        
                                        treev.insert("", 'end', text ="L7",
                                                values =(amount,expense), tags = ("even",))
                                if c%2 != 0:                        
                                        treev.insert("", 'end', text ="L7",
                                                values =(amount,expense), tags = ("odd",))
                                                        
                        treev.tag_configure('even', background='#e1dddd', font=font22) 
                        treev.tag_configure('odd', background='#f5f3f3', font=font22) 

        def clientsExpenses():
                treev = ttk.Treeview(self, selectmode ='browse', style="mystyle.Treeview", height=height)
                treev.pack()
                style = ttk.Style()
                style.configure("mystyle.Treeview", background = baseColor ,rowheight=100,
                highlightthickness=0, bd=0, font=('Helvetica', 14)) 
                style.configure("mystyle.Treeview.Heading", font=font20) 
                style.layout("mystyle.Treeview", [('mystyle.Treeview.treearea', {'sticky': 'nswe'})]) 
                verscrlbar = ttk.Scrollbar(self,
                                orient ="vertical",
                                command = treev.yview)
                verscrlbar.pack(side ='left', fill ='x')   
                treev.configure(xscrollcommand = verscrlbar.set)
                treev["columns"] = ("1", "2", "3")
                treev['show'] = 'headings'
                treev.column("1", width = int(width*0.42),anchor ='e')
                treev.column("2", width = int(width*0.55),anchor ='e')
                treev.column("3", width = int(width*0.14),anchor ='e')
                treev.heading("1", text ="اجمالى المصروفات")
                treev.heading("2", text ="الاسم")
                treev.heading("3", text ="الكود")
                c = 0
                for expenses_sheet in wb0.worksheets:
                        c += 1
                        client_expenses = 0
                        ws = expenses_sheet
                        if (ws.cell(row=1,column=1).value) != None or expenses_sheet.title != 'مصروفات اداريه' or expenses_sheet.title != 'الخدمات':
                                for i in range(4, ws.max_row+1):
                                        try:
                                                celldate = date(
                                                        ws.cell(row=i, column=4).value,
                                                        ws.cell(row=i, column=5).value,
                                                        ws.cell(row=i, column=6).value)
                                        except:
                                                continue
                                        if celldate >= startdate.get_date() and celldate <= enddate.get_date():
                                                        if ws.cell(row=i, column=2).value != None:
                                                                client_expenses += ws.cell(row=i, column=2).value 
                                if c%2 == 0:
                                        treev.insert("", 'end', text ="L7",
                                                        values =(
                                                        '-' if int(client_expenses) == None else str(client_expenses),
                                                        '-' if (ws.cell(row=1,column=1).value) == None else str(ws.cell(row=1,column=1).value),
                                                        '-' if (ws.cell(row=1,column=2).value) == None else str(ws.cell(row=1,column=2).value),
                                                        ), tags = ('even',))
                                if c%2 != 0: 
                                        treev.insert("", 'end', text ="L7",
                                                        values =(
                                                        '-' if int(client_expenses) == None else str(client_expenses),
                                                        '-' if (ws.cell(row=1,column=1).value) == None else str(ws.cell(row=1,column=1).value),
                                                        '-' if (ws.cell(row=1,column=2).value) == None else str(ws.cell(row=1,column=2).value),
                                                        ), tags = ('odd',))                       
                        treev.tag_configure('even', background='#e1dddd', font=font22) 
                        treev.tag_configure('odd', background='#f5f3f3', font=font22) 
                
                
        Button(self, height = 1, width = 15, bg = '#05659E', fg = 'white',
        activebackground='#43516C', font = font20, bd = '4px solid #DBA531', 
                text='المصروفات الاداريه',
                command=viewExpenses).place(relx=.5, rely=.55,anchor= CENTER) 

        Button(self, height = 2, width = 20, bg = '#D85426', fg = 'white',
        activebackground='#D85426', font = 'fantasy 24 bold', bd = '8px solid #DBA531', 
                text='مصروفات العملاء',
                command= clientsExpenses).place(relx=.5, rely=.2,anchor= CENTER)                                                                     

        Button(self, height = 1, width = 8, bg = 'grey', fg = 'white',
        activebackground='#43516C', font = 'fantasy 15 bold', bd = '4px solid #DBA531', 
                text='اغلاق',
                command=self.destroy).place(relx=.1, rely=.95,anchor= CENTER)


# -----------------------------------------------------------------------------------------------------------------------------

class Expenses_Form(tk.Toplevel):
      def __init__(self, parent):
        super().__init__(parent)

        self.geometry(f'{width}x{height}')
        self.title('المصروفات')
        self['bg']=baseColor
        
        entries = []
        for sheet in wb.worksheets:
                if sheet.cell(row=1,column=1).value != None or sheet.title != 'مصروفات اداريه' or sheet.title != 'الخدمات':
                        entries.append(f'{str(sheet.cell(row=1,column=1).value)} ({str(sheet.cell(row=1,column=2).value)})')                  

        Label(self, text='اسم العميل',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.75, rely=.1,anchor= CENTER, width=int(width*0.14))

        def select_input(event):
                value = event.widget.get()
                if value == '':
                        menu['values'] = entries
                else:
                        data = []
                        for item in entries:
                                if value.lower() in item.lower():
                                        data.append(item)

                        menu['values'] = data

        name = StringVar()
        menu = ttk.Combobox(self, textvariable= name, font=font20,values = entries )
        app.option_add('*TCombobox*Listbox.font', font20)
        menu.place(relx=.5, rely=.1,anchor= CENTER, width=int(width*0.4))
        menu.bind('<KeyRelease>', select_input)

        Label(self, text='.اترك خانة (اسم العميل) فارغه اذا كان المصروف غير متعلق بعميل معين',
        bg = baseColor, fg = labelColor ,font = font10).place(relx=.59, rely=.15,anchor= CENTER)


        Label(self, text='المبلغ',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.75, rely=.25,anchor= CENTER, width=int(width*0.14))

        amount = IntVar()
        entry1 = ttk.Entry(self, textvariable = amount, justify = LEFT, font = font20)
        entry1.place(relx=.5, rely=.25,anchor= CENTER, width=int(width*0.4), height=50)

        Label(self, text='.برجاء عدم ترك الخانه فارغه',
        bg = baseColor, fg = labelColor ,font = font10).place(relx=.65, rely=.31,anchor= CENTER)
        

        Label(self, text='المصروف',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.75, rely=.4,anchor= CENTER, width=int(width*0.14))

        comment = StringVar()
        entry1 = ttk.Entry(self, textvariable=comment, justify = LEFT, font = font20)
        entry1.place(relx=.5, rely=.4,anchor= CENTER, width=int(width*0.4), height=50)

        Label(self, text='.برجاء عدم ترك الخانه فارغه',
        bg = baseColor, fg = labelColor ,font = font10).place(relx=.65, rely=.46,anchor= CENTER)


        Label(self, text='التاريخ',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.75, rely=.55,anchor= CENTER, width=int(width*0.14))

        c_date = DateEntry(self,width=30,bg="darkblue",fg="white",year=datetime.now().year,
        month=datetime.now().month,day=datetime.now().day
        ,locale='en_US', date_pattern='dd/MM/yyyy',font = font18)
        c_date.place(relx=.625, rely=.55,anchor= CENTER, width=int(width*0.14), height=50)
        
        
        def addExpenses():
                try:
                        amount.get()
                except:
                        messagebox.showerror('Invalid!','من فضلك قم بادخال الخانات بطريقه صحيحه ') 
                        print('invalid amount')
                        return False 

                if amount.get() == '':
                        messagebox.showerror('Invalid!','من فضلك قم بادخال الخانات المطلوبه   ')
                        print('No amount')
                        return False 

                if comment.get() == '':
                        messagebox.showerror('Invalid!','من فضلك قم بادخال خانة المصروف ')
                        print('No Expense')
                        return False     

                if amount.get() == 0:
                        messagebox.showerror('Invalid!','من فضلك قم بتعديل خانة المبلغ')
                        print('amount is 0')
                        return False  

                else:                
                        if name.get() == '':
                                try:
                                        ws = wb0['مصروفات اداريه']
                                except:
                                        ws = wb0.create_sheet('مصروفات اداريه') 
                                        ws.append(['التاريخ', 'المبلغ','المصروف'])
                                        ws.append([c_date.get_date().strftime("%d/%m/%Y")
                                                                ,amount.get()
                                                                ,comment.get()
                                                                ,c_date.get_date().year
                                                                ,c_date.get_date().month
                                                                ,c_date.get_date().day,])
                                        return True                        

                                inserted_date = date(c_date.get_date().year,c_date.get_date().month, c_date.get_date().day)
                                # check if transaction needs to be sorted by comparing date
                                exists_dates = []
                                for i in range (2, ws.max_row+1):
                                        try:
                                                exists_dates.append(date(
                                                ws.cell(column=4, row=i).value,
                                                ws.cell(column=5, row=i).value,
                                                ws.cell(column=6, row=i).value,))
                                        except:
                                                continue
                                flag = False
                                for cell_date in exists_dates:
                                        if inserted_date < cell_date:
                                # add all transactions to rows_list
                                                rows_list = []
                                                for i in range (2, ws.max_row+1):
                                                        try:
                                                                cur_row_date = date(
                                                                ws.cell(column=4, row=i).value,
                                                                ws.cell(column=5, row=i).value,
                                                                ws.cell(column=6, row=i).value,)
                                                        except:
                                                                continue        
                                                        cur_row = [ws.cell(column=1, row=i).value,
                                                        ws.cell(column=2, row=i).value,
                                                        ws.cell(column=3, row=i).value,
                                                        ws.cell(column=4, row=i).value,
                                                        ws.cell(column=5, row=i).value,
                                                        ws.cell(column=6, row=i).value,
                                                        cur_row_date] 
                                                        rows_list.append(cur_row)
                                                
                                                rows_list.append([c_date.get_date().strftime("%d/%m/%Y")
                                                        ,amount.get()
                                                        ,comment.get()
                                                        ,c_date.get_date().year
                                                        ,c_date.get_date().month
                                                        ,c_date.get_date().day, inserted_date]) 
                                
                                                rows_list.sort(key=lambda x : x[6])
                                                # delete rows 
                                                ws.delete_rows(2, ws.max_row)
                                                #add sorted rows
                                                ws.append(rows_list[0])
                                                rows_list.pop(0)
                                                for row in rows_list:
                                                        ws.append(row)

                                                wb0.save(file0)
                                                flag = True
                                                self.destroy()
                                                messagebox.showinfo('Done','تم الحفظ بنجاح ')
                                                break

                                if flag == False:
                                        ws.append([c_date.get_date().strftime("%d/%m/%Y")
                                                ,amount.get()
                                                ,comment.get()
                                                ,c_date.get_date().year
                                                ,c_date.get_date().month
                                                ,c_date.get_date().day, inserted_date]) 
                                        wb0.save(file0)
                                        self.destroy()
                                        messagebox.showinfo('Done','تم الحفظ بنجاح ')  
                        else:        
                                try:
                                        ws = wb0[name.get()]
                                except:
                                        messagebox.showerror('Not Exists!','الاسم غير موجود') 
                                        print('clients expenes file not found')
                                        return False

                                inserted_date = date(c_date.get_date().year,c_date.get_date().month, c_date.get_date().day)
                                # first entry
                                if ws.cell(row=4, column=1).value == None or ws.cell(row=4, column=1).value == '' :
                                        ws['A4'] = c_date.get_date().strftime("%d/%m/%Y")
                                        ws['B4'] = amount.get()
                                        ws['C4'] = comment.get()
                                        ws['D4'] = c_date.get_date().year
                                        ws['E4'] = c_date.get_date().month
                                        ws['F4'] = c_date.get_date().day
                                        ws['G4'] = inserted_date
                                        wb0.save(file0)
                                        self.destroy()
                                        messagebox.showinfo('Done','تم الحفظ بنجاح ')
                                #no first entry        
                                else:        
                                        # check if transaction needs to be sorted by comparing date
                                        exists_dates = []
                                        for i in range (4, ws.max_row+1):
                                                try:
                                                        exists_dates.append(date(
                                                        ws.cell(column=4, row=i).value,
                                                        ws.cell(column=5, row=i).value,
                                                        ws.cell(column=6, row=i).value,))
                                                except:
                                                        continue

                                        flag = False
                                        for cell_date in exists_dates:
                                                if inserted_date < cell_date:
                                        # add all transactions to rows_list
                                                        rows_list = []
                                                        for i in range (4, ws.max_row+1):
                                                                try:
                                                                        cur_row_date = date(
                                                                        ws.cell(column=4, row=i).value,
                                                                        ws.cell(column=5, row=i).value,
                                                                        ws.cell(column=6, row=i).value,)
                                                                except:
                                                                        continue        
                                                                cur_row = [ws.cell(column=1, row=i).value,
                                                                ws.cell(column=2, row=i).value,
                                                                ws.cell(column=3, row=i).value,
                                                                ws.cell(column=4, row=i).value,
                                                                ws.cell(column=5, row=i).value,
                                                                ws.cell(column=6, row=i).value,
                                                                cur_row_date] 
                                                                rows_list.append(cur_row)
                                                        
                                                        rows_list.append([c_date.get_date().strftime("%d/%m/%Y")
                                                                ,amount.get()
                                                                ,comment.get()
                                                                ,c_date.get_date().year
                                                                ,c_date.get_date().month
                                                                ,c_date.get_date().day, inserted_date]) 
                                        
                                                        rows_list.sort(key=lambda x : x[6])
                                                        # delete rows 
                                                        ws.delete_rows(4, ws.max_row)
                                                        #add sorted rows
                                                        ws.append(rows_list[0])
                                                        rows_list.pop(0)
                                                        for row in rows_list:
                                                                ws.append(row)

                                                        wb0.save(file0)
                                                        flag = True
                                                        # self.destroy()
                                                        messagebox.showinfo('Done','تم الحفظ بنجاح ')
                                                        break
                                        if flag == False:
                                                ws.append([c_date.get_date().strftime("%d/%m/%Y")
                                                        ,amount.get()
                                                        ,comment.get()
                                                        ,c_date.get_date().year
                                                        ,c_date.get_date().month
                                                        ,c_date.get_date().day, inserted_date]) 
                                                wb0.save(file0)
                                                # self.destroy()
                                                messagebox.showinfo('Done','تم الحفظ بنجاح ')


        Button(self, height = 1, width = 10, bg = '#05659E', fg = 'white',
        activebackground='#43516C', font = font20, bd = '4px solid #DBA531', 
                text='حفظ',
                command=addExpenses).place(relx=.5, rely=.9,anchor= CENTER)
        
        Button(self, height = 1, width = 8, bg = 'grey', fg = 'white',
        activebackground='#43516C', font = 'fantasy 15 bold', bd = '4px solid #DBA531', 
                text='اغلاق',
                command=self.destroy).place(relx=.1, rely=.95,anchor= CENTER)        
        
# -----------------------------------------------------------------------------------------------------------------------------

class Revenues_View(tk.Toplevel):
      def __init__(self, parent):
        super().__init__(parent)

        self.geometry(f'{width}x{height}')
        self.title('الايرادات')
        self['bg']=baseColor


        Label(self, text='من',
        bg = '#43516C', fg = 'white', font = font20,
        borderwidth=5, relief="ridge", padx=20, pady=10).place(relx=.7, rely=.4,anchor= CENTER, width=int(width*0.28))

        startdate = DateEntry(self,width=30,bg="darkblue",fg="white", font = font18,
        year=datetime.now().year,month=datetime.now().month,day=datetime.now().day
        ,locale='en_US', date_pattern='dd/MM/yyyy')
        startdate.place(relx=.7, rely=.5,anchor= CENTER, width=int(width*0.14), height=50)
        

        Label(self, text='الى',
        bg = '#43516C', fg = 'white', font = font20,
        borderwidth=5, relief="ridge", padx=20, pady=10).place(relx=.3, rely=.4,anchor= CENTER, width=int(width*0.28))

        enddate = DateEntry(self,width=30,bg="darkblue",fg="white", font = font18,
        year=datetime.now().year,month=datetime.now().month,day=datetime.now().day
        ,locale='en_US', date_pattern='dd/MM/yyyy')
        enddate.place(relx=.3, rely=.5,anchor= CENTER, width=int(width*0.14), height=50)
        

        def clientsRevenues():
                treev = ttk.Treeview(self, selectmode ='browse', style="mystyle.Treeview", height=height)
                treev.pack()
                style = ttk.Style()
                style.configure("mystyle.Treeview", background = baseColor ,rowheight=100,
                highlightthickness=0, bd=0, font=('Helvetica', 14)) 
                style.configure("mystyle.Treeview.Heading", font=font20) 
                style.layout("mystyle.Treeview", [('mystyle.Treeview.treearea', {'sticky': 'nswe'})]) 
                verscrlbar = ttk.Scrollbar(self,
                                orient ="vertical",
                                command = treev.yview)
                verscrlbar.pack(side ='left', fill ='x')   
                treev.configure(xscrollcommand = verscrlbar.set)
                treev["columns"] = ("1", "2", "3", "4", "5")
                treev['show'] = 'headings'
                treev.column("1", width = int(width*0.15),anchor ='e')
                treev.column("2", width = int(width*0.15),anchor ='e')
                treev.column("3", width = int(width*0.15),anchor ='e')
                treev.column("4", width = int(width*0.4),anchor ='e')
                treev.column("5", width = int(width*0.15),anchor ='e')
                treev.heading("1", text ="صافى الربح")
                treev.heading("2", text ="المصروف")
                treev.heading("3", text ="الايراد")
                treev.heading("4", text ="الاسم")
                treev.heading("5", text ="الكود")

                c = 0
                for client_sheet in wb.worksheets:
                        client_expenses = 0
                        client_revenue = 0
                        amount = 0
                        ws = client_sheet
                        c += 1
                        if client_sheet.title != 'الخدمات' or client_sheet.title != 'المصروفات الاداريه':
                                for i in range(4, ws.max_row+1):
                                        try:
                                                celldate = date(
                                                        ws.cell(row=i, column=7).value,
                                                        ws.cell(row=i, column=8).value,
                                                        ws.cell(row=i, column=9).value)
                                        except:
                                                continue
                                        if celldate >= startdate.get_date() and celldate <= enddate.get_date():
                                                if ws.cell(row=i, column=4).value != None:
                                                        client_revenue += ws.cell(row=i, column=4).value

                                for expenses_sheet in wb0.worksheets:
                                        if expenses_sheet.title == client_sheet.title: 
                                                ws = expenses_sheet
                                                for i in range(4, ws.max_row+1):
                                                        try:
                                                                celldate = date(
                                                                        ws.cell(row=i, column=4).value,
                                                                        ws.cell(row=i, column=5).value,
                                                                        ws.cell(row=i, column=6).value)
                                                        except:
                                                                continue
                                                        if celldate >= startdate.get_date() and celldate <= enddate.get_date():
                                                                if ws.cell(row=i, column=2).value != None:
                                                                        client_expenses += ws.cell(row=i, column=2).value

                                amount = client_revenue - client_expenses
                                if c%2 == 0:
                                        treev.insert("", 'end', text ="L7",
                                                        values =(
                                                        '-' if int(amount) == None else str(amount),
                                                        '-' if int(client_expenses) == None else str(client_expenses),
                                                        '-' if int(client_revenue) == None else str(client_revenue),
                                                        '-' if (ws.cell(row=1,column=1).value) == None else str(ws.cell(row=1,column=1).value),
                                                        '-' if (ws.cell(row=1,column=2).value) == None else str(ws.cell(row=1,column=2).value),
                                                        ), tags = ('even',))
                                if c%2 != 0:
                                        treev.insert("", 'end', text ="L7",
                                                        values =(
                                                        '-' if int(amount) == None else str(amount),
                                                        '-' if int(client_expenses) == None else str(client_expenses),
                                                        '-' if int(client_revenue) == None else str(client_revenue),
                                                        '-' if (ws.cell(row=1,column=1).value) == None else str(ws.cell(row=1,column=1).value),
                                                        '-' if (ws.cell(row=1,column=2).value) == None else str(ws.cell(row=1,column=2).value),
                                                        ), tags = ('odd',))                        
                treev.tag_configure('even', background='#e1dddd', font=font22) 
                treev.tag_configure('odd', background='#f5f3f3', font=font22) 


        def totalRevenues():      
                total_revenue = 0
                total_expenses = 0
                income = 0
                for sheet in wb.worksheets:
                        if sheet.title != 'الخدمات':
                                ws = sheet
                                for i in range(4, ws.max_row+1):
                                        try:
                                                celldate = date(
                                                        ws.cell(row=i, column=7).value,
                                                        ws.cell(row=i, column=8).value,
                                                        ws.cell(row=i, column=9).value)
                                        except:
                                                continue
                                        if celldate >= startdate.get_date() and celldate <= enddate.get_date():
                                                try:
                                                        total_revenue += ws.cell(row=i, column=4).value
                                                except:
                                                        continue 
                                                
                        for expenses_sheet in wb0.worksheets:
                                ws = expenses_sheet
                                for i in range(0, ws.max_row+1):
                                        try:    
                                                celldate = date(
                                                        ws.cell(row=i, column=4).value,
                                                        ws.cell(row=i, column=5).value,
                                                        ws.cell(row=i, column=6).value)
                                        except:
                                                continue
                                        if celldate >= startdate.get_date() and celldate <= enddate.get_date():
                                                try:
                                                        total_expenses += ws.cell(row=i, column=2).value
                                                except:
                                                        continue
                                                                
                        income = total_revenue - total_expenses 

                        Label(self, text='الايرادات',
                        bg = baseColor, fg = labelColor ,font = font20).place(relx=.8, rely=.65,anchor= CENTER, width=int(width*0.14))
                        Label(self, text='المصروفات',
                        bg = baseColor, fg = labelColor ,font = font20).place(relx=.5, rely=.65,anchor= CENTER, width=int(width*0.14))
                        Label(self, text='صافى الربح',
                        bg = baseColor, fg = labelColor ,font = font20).place(relx=.2, rely=.65,anchor= CENTER, width=int(width*0.14))
                        Label(self, text=str(total_revenue),
                        bg = '#E85662', fg = 'white', font = 'fantasy 30 bold',
                        borderwidth=5, relief="ridge", padx=20, pady=10).place(relx=.8, rely=.75,anchor= CENTER, width=int(width*0.16))
                        Label(self, text='_',
                        bg = baseColor, fg = labelColor ,font = 'fantasy 30 bold').place(relx=.65, rely=.72,anchor= CENTER, width=int(width*0.14))
                        Label(self, text=str(total_expenses),
                        bg = '#E85662', fg = 'white', font = 'fantasy 30 bold',
                        borderwidth=5, relief="ridge", padx=20, pady=10).place(relx=.5, rely=.75,anchor= CENTER, width=int(width*0.16))
                        Label(self, text='=',
                        bg = baseColor, fg = labelColor ,font = 'fantasy 30 bold').place(relx=.35, rely=.75,anchor= CENTER, width=int(width*0.14))
                        Label(self, text=str(income),
                        bg = '#E85662', fg = 'white', font = 'fantasy 30 bold',
                        borderwidth=5, relief="ridge", padx=20, pady=10).place(relx=.2, rely=.75,anchor= CENTER, width=int(width*0.16))
                                        

        Button(self, height = 1, width = 17, bg = '#05659E', fg = 'white',
        activebackground='#43516C', font = font20, bd = '4px solid #DBA531', 
                text='اجمالى الايرادات و المصروفات',
                command=totalRevenues).place(relx=.5, rely=.55,anchor= CENTER)        

        Button(self, height = 2, width = 20, bg = '#D85426', fg = 'white',
        activebackground='#D85426', font = 'fantasy 24 bold', bd = '8px solid #DBA531', 
                text='تفاصيل الايرادات',
                command= clientsRevenues).place(relx=.5, rely=.2,anchor= CENTER) 

        Button(self, height = 1, width = 8, bg = 'grey', fg = 'white',
        activebackground='#43516C', font = 'fantasy 15 bold', bd = '4px solid #DBA531', 
                text='اغلاق',
                command=self.destroy).place(relx=.1, rely=.95,anchor= CENTER)

        
# -----------------------------------------------------------------------------------------------------------------------------

class Payment_Form(tk.Toplevel):
      def __init__(self, parent):
        super().__init__(parent)

        self.geometry(f'{width}x{height}')
        self.title('المدفوعات')
        self['bg']=baseColor


        entries = []
        for sheet in wb.worksheets:
                if sheet.cell(row=1,column=1).value != None or sheet.title != 'الخدمات':
                        entries.append(f'{str(sheet.cell(row=1,column=1).value)} ({str(sheet.cell(row=1,column=2).value)})')                


        Label(self, text='الاسم',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.8, rely=.1,anchor= CENTER, width=int(width*0.14))

        def select_input(event):
                value = event.widget.get()
                if value == '':
                        menu['values'] = entries
                else:
                        data = []
                        for item in entries:
                                if value.lower() in item.lower():
                                        data.append(item)

                        menu['values'] = data

        name = StringVar()
        menu = ttk.Combobox(self, textvariable= name, font=font20,values = entries )
        app.option_add('*TCombobox*Listbox.font', font20)
        menu.place(relx=.55, rely=.1,anchor= CENTER, width=int(width*0.4))
        menu.bind('<KeyRelease>', select_input)

        Label(self, text='.برجاء عدم ترك الخانه فارغه',
        bg = baseColor, fg = labelColor ,font = font10).place(relx=.705, rely=.15,anchor= CENTER)


        Label(self, text='المدفوع',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.8, rely=.25,anchor= CENTER, width=int(width*0.14))

        amount = IntVar()
        entry1 = ttk.Entry(self, textvariable = amount, justify = LEFT, font = font20)
        entry1.place(relx=.55, rely=.25,anchor= CENTER, width=int(width*0.4), height=50)

        Label(self, text='.برجاء عدم ترك الخانه فارغه',
        bg = baseColor, fg = labelColor ,font = font10).place(relx=.705, rely=.31,anchor= CENTER)
        

        Label(self, text='ملاحظات',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.8, rely=.4,anchor= CENTER, width=int(width*0.14))

        comment =  Text(self, font = font20, height=3)
        comment.place(relx=.55, rely=.42,anchor= CENTER, width=int(width*0.4))


        Label(self, text='التاريخ',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.8, rely=.55,anchor= CENTER, width=int(width*0.14))

        c_date = DateEntry(self,width=30,bg="darkblue",fg="white",year=datetime.now().year,
        month=datetime.now().month,day=datetime.now().day
        ,locale='en_US', date_pattern='dd/MM/yyyy',font = font18)
        c_date.place(relx=.675, rely=.55,anchor= CENTER, width=int(width*0.14), height=50)
        
        
        def payment():
                try:
                        amount.get() and name.get()
                except:
                        messagebox.showerror('Invalid!','من فضلك قم بادخال الخانات بطريقه صحيحه ') 
                        print('invalid name or amount')
                        return False 

                if amount.get() == '' or amount.get() == None :
                        messagebox.showerror('Invalid!','من فضلك قم بادخال خانة المدفوع بطريقه صحيحه ')
                        print('no amount')
                        return False

                if name.get() == '':
                        messagebox.showerror('Invalid!','من فضلك قم بادخال خانة الاسم بطريقه صحيحه ')
                        # self.destroy()
                        return        

                # for sheet in wb.worksheets:
                try:
                        ws = wb[name.get()]
                except:
                        messagebox.showerror('Not Exists!','الاسم غير موجود') 
                        print('Clients sheet not exists')
                        return False
                        # self.destroy()

                else:   
                        finalAmount = ws.cell(column=5, row=ws.max_row).value
                        if type(finalAmount) == str or finalAmount == 0 or finalAmount == None:
                                finalAmount = 0  
                        
                        inserted_date = date(c_date.get_date().year,c_date.get_date().month, c_date.get_date().day)
                        # check if transaction needs to be sorted by comparing date
                        exists_dates = []
                        for i in range (4, ws.max_row+1):
                                try:
                                        exists_dates.append(date(
                                        ws.cell(column=7, row=i).value,
                                        ws.cell(column=8, row=i).value,
                                        ws.cell(column=9, row=i).value,))
                                except:
                                        continue
                        flag = False        
                        # if transaction needs to be sorted
                        for cell_date in exists_dates:
                                if inserted_date < cell_date:
                                # add all transactions to rows_list
                                        rows_list = []
                                        for i in range (4, ws.max_row+1):
                                                try:
                                                        cur_row_date = date(
                                                                ws.cell(column=7, row=i).value,
                                                                ws.cell(column=8, row=i).value,
                                                                ws.cell(column=9, row=i).value,)
                                                except:
                                                        continue        
                                                cur_row = [    
                                                ws.cell(column=1, row=i).value,
                                                ws.cell(column=2, row=i).value,
                                                ws.cell(column=3, row=i).value,
                                                ws.cell(column=4, row=i).value,
                                                ws.cell(column=5, row=i).value,
                                                ws.cell(column=6, row=i).value,
                                                ws.cell(column=7, row=i).value,
                                                ws.cell(column=8, row=i).value,
                                                ws.cell(column=9, row=i).value,
                                                cur_row_date]

                                                rows_list.append(cur_row)
                                        
                                        rows_list.append([c_date.get_date().strftime("%d/%m/%Y"),
                                                '-',0,
                                                amount.get(),
                                                finalAmount - amount.get(),
                                                comment.get("1.0",END),
                                                c_date.get_date().year, c_date.get_date().month, c_date.get_date().day, inserted_date]) 
                                        rows_list.sort(key=lambda x : x[9])
                                        # delete rows 
                                        ws.delete_rows(4, ws.max_row)
                                        #add sorted rows
                                        ws.append([rows_list[0][0],rows_list[0][1],rows_list[0][2],rows_list[0][3],int(rows_list[0][2])-int(rows_list[0][3]),rows_list[0][5],rows_list[0][6],rows_list[0][7],rows_list[0][8],rows_list[0][9]])
                                        
                                        rows_list.pop(0)
                                        for row in rows_list:
                                                try:
                                                        ws.append([row[0],row[1],row[2],row[3],int(row[2])-int(row[3])+int(ws.cell(column=5,row=ws.max_row).value),row[5],row[6],row[7],row[8],row[9]])
                                                except:
                                                        continue
                                        wb.save(file)
                                        flag = True
                                        self.destroy()
                                        messagebox.showinfo('Done','تم الحفظ بنجاح ')
                                        break

                        if flag == False:
                                finalAmount = ws.cell(column=5,row=int(ws.max_row)).value
                                if type(finalAmount) == str or finalAmount == 0 or finalAmount == None:
                                        finalAmount = 0       
                                ws.append([c_date.get_date().strftime("%d/%m/%Y"),
                                        '-', 0,
                                        amount.get(),
                                        finalAmount - amount.get(), 
                                        comment.get("1.0",END),
                                        c_date.get_date().year, c_date.get_date().month, c_date.get_date().day, inserted_date])
                                wb.save(file)
                                self.destroy()
                                messagebox.showinfo('Done','تم الحفظ بنجاح ')


        Button(self, height = 1, width = 10, bg = '#05659E', fg = 'white',
        activebackground='#43516C', font = font20, bd = '4px solid #DBA531', 
                text='حفظ',
                command=payment).place(relx=.5, rely=.9,anchor= CENTER)
        
        Button(self, height = 1, width = 8, bg = 'grey', fg = 'white',
        activebackground='#43516C', font = 'fantasy 15 bold', bd = '4px solid #DBA531', 
                text='اغلاق',
                command=self.destroy).place(relx=.1, rely=.95,anchor= CENTER)     
        
# # --------------------------------------------------------------------------------------------------------------------------------------------                

class Search(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)

        self.geometry(f'{width}x{height}')
        self.title('البحث عن عميل')
        self['bg']=baseColor

        entries = []
        for sheet in wb.worksheets:
                if sheet.cell(row=1,column=1).value != None or sheet.title != 'الخدمات':
                        entries.append(f'{str(sheet.cell(row=1,column=1).value)} ({str(sheet.cell(row=1,column=2).value)})')


        def search_input(event):
                value = event.widget.get()
                if value == '':
                        menu['values'] = entries
                else:
                        data = []
                        for item in entries:
                                if value.lower() in item.lower():
                                        data.append(item)

                        menu['values'] = data

        
        Label(self, text='اسم العميل',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.69, rely=.35,anchor= CENTER, width=int(width*0.14))

        name = StringVar()
        menu = ttk.Combobox(self, textvariable= name, font=font20,values = entries )
        app.option_add('*TCombobox*Listbox.font', font20)
        menu.place(relx=.5, rely=.4,anchor= CENTER, width=int(width*0.47))
        menu.bind('<KeyRelease>', search_input)

        Label(self, text='.برجاء عدم ترك الخانه فارغه',
        bg = baseColor, fg = labelColor ,font = font10).place(relx=.69, rely=.45,anchor= CENTER)


        def search():
                treev = ttk.Treeview(self, selectmode ='browse', style="mystyle.Treeview", height=height)
                treev.pack()
                style = ttk.Style()
                style.configure("mystyle.Treeview", background = baseColor ,rowheight=100,
                highlightthickness=0, bd=0, font=('Helvetica', 14)) 
                style.configure("mystyle.Treeview.Heading", font=font20) 
                style.layout("mystyle.Treeview", [('mystyle.Treeview.treearea', {'sticky': 'nswe'})]) 
                verscrlbar = ttk.Scrollbar(self,
                                orient ="vertical",
                                command = treev.yview)
                verscrlbar.pack(side ='left', fill ='x')   
                treev.configure(xscrollcommand = verscrlbar.set)
                treev["columns"] = ("1", "2", "3", "4", "5", "6")
                treev['show'] = 'headings'
                treev.column("6", width = int(width*0.15),anchor ='e')
                treev.column("5", width = int(width*0.3),anchor ='e')
                treev.column("4", width = int(width*0.1),anchor ='e')
                treev.column("3", width = int(width*0.1),anchor ='e')
                treev.column("2", width = int(width*0.1),anchor ='e')
                treev.column("1", width = int(width*0.35),anchor ='e')
                treev.heading("6", text ="التاريخ")
                treev.heading("5", text ="الخدمه")
                treev.heading("4", text ="التكلفه")
                treev.heading("3", text ="المدفوع")
                treev.heading("2", text ="الرصيد")
                treev.heading("1", text ="الملاحظات")

        
                try:
                        ws = wb[name.get()]
                except:
                        messagebox.showerror('Not Exists!','الاسم غير موجود') 
                        print('Clients sheet do not exists')
                        return False       
                        
                for i in range (4, ws.max_row+1):
                        if (ws.cell(row=i,column=1).value) != None:
                                if i%2 == 0:
                                        treev.insert("", 'end', text =i,
                                                        values =(
                                                        '-' if (ws.cell(row=i,column=6).value) == None else str(ws.cell(row=i,column=6).value), 
                                                        '-' if (ws.cell(row=i,column=5).value) == None else str(ws.cell(row=i,column=5).value),
                                                        '-' if (ws.cell(row=i,column=4).value) == None else str(ws.cell(row=i,column=4).value),
                                                        '-' if (ws.cell(row=i,column=3).value) == None else str(ws.cell(row=i,column=3).value),
                                                        '-' if (ws.cell(row=i,column=2).value) == None else str(ws.cell(row=i,column=2).value),
                                                        '-' if (ws.cell(row=i,column=1).value) == None else str(ws.cell(row=i,column=1).value)), tags = ('even',))
                                else:
                                        treev.insert("", 'end', text =i,
                                                        values =(
                                                        '-' if (ws.cell(row=i,column=6).value) == None else str(ws.cell(row=i,column=6).value), 
                                                        '-' if (ws.cell(row=i,column=5).value) == None else str(ws.cell(row=i,column=5).value),
                                                        '-' if (ws.cell(row=i,column=4).value) == None else str(ws.cell(row=i,column=4).value),
                                                        '-' if (ws.cell(row=i,column=3).value) == None else str(ws.cell(row=i,column=3).value),
                                                        '-' if (ws.cell(row=i,column=2).value) == None else str(ws.cell(row=i,column=2).value),
                                                        '-' if (ws.cell(row=i,column=1).value) == None else str(ws.cell(row=i,column=1).value)), tags = ('odd',))                        
                        
                treev.tag_configure('odd', background='#e1dddd', font=font22) 
                treev.tag_configure('even', background= '#f5f3f3', font=font22)

                def OnDelete(event):
                        response = messagebox.askquestion('Delete!', f'هل انت متأكد من حذف هذه المعامله؟')
                        if response == 'yes' :
                                item = treev.selection()[0]
                                ws.delete_rows(treev.item(item,"text"),1)
                                treev.delete(item)
                                wb.save(file)
                                rows_list = []
                                for i in range (4, ws.max_row+1):
                                        try:
                                                cur_row_date = date(
                                                        ws.cell(column=7, row=i).value,
                                                        ws.cell(column=8, row=i).value,
                                                        ws.cell(column=9, row=i).value,)
                                        except:
                                                continue        
                                        cur_row = [    
                                        ws.cell(column=1, row=i).value,
                                        ws.cell(column=2, row=i).value,
                                        ws.cell(column=3, row=i).value,
                                        ws.cell(column=4, row=i).value,
                                        ws.cell(column=5, row=i).value,
                                        ws.cell(column=6, row=i).value,
                                        ws.cell(column=7, row=i).value,
                                        ws.cell(column=8, row=i).value,
                                        ws.cell(column=9, row=i).value,
                                        cur_row_date]

                                        rows_list.append(cur_row)
                                # sort rows_list by date
                                if len(rows_list) > 0:
                                        rows_list.sort(key=lambda x : x[9])
                                        # delete rows 
                                        ws.delete_rows(4, ws.max_row)
                                        #add sorted rows
                                        ws.append([rows_list[0][0],rows_list[0][1],rows_list[0][2],rows_list[0][3],int(rows_list[0][2])-int(rows_list[0][3]),rows_list[0][5],rows_list[0][6],rows_list[0][7],rows_list[0][8],rows_list[0][9]])
                                        rows_list.pop(0)
                                        for row in rows_list:
                                                try:
                                                        ws.append([row[0],row[1],row[2],row[3],row[2]-row[3]+int(ws.cell(column=5,row=ws.max_row).value),row[5],row[6],row[7],row[8],row[9],])
                                                except:
                                                        continue
                                        wb.save(file)

                                self.destroy()
                                messagebox.showinfo('Done!','تم الحذف')
                        else:
                                return False      
                                
                treev.bind("<d>", OnDelete)
                self.title( name.get())
                

        def clients():
                treev = ttk.Treeview(self, selectmode ='browse', style="mystyle.Treeview", height=height)
                treev.pack()
                style = ttk.Style()
                style.configure("mystyle.Treeview", background = baseColor ,rowheight=100,
                highlightthickness=0, bd=0, font=('Helvetica', 14)) 
                style.configure("mystyle.Treeview.Heading", font=font20) 
                style.layout("mystyle.Treeview", [('mystyle.Treeview.treearea', {'sticky': 'nswe'})]) 
                verscrlbar = ttk.Scrollbar(self,
                                orient ="vertical",
                                command = treev.yview)
                verscrlbar.pack(side ='left', fill ='x')   
                treev.configure(xscrollcommand = verscrlbar.set)
                treev["columns"] = ("1", "2", "3", "4", "5", "6")
                treev['show'] = 'headings'
                treev.column("1", width = int(width*0.2),anchor ='e')
                treev.column("2", width = int(width*0.1),anchor ='e')
                treev.column("3", width = int(width*0.1),anchor ='e')
                treev.column("4", width = int(width*0.1),anchor ='e')
                treev.column("5", width = int(width*0.4),anchor ='e')
                treev.column("6", width = int(width*0.1),anchor ='e')
                treev.heading("1", text ="العنوان")
                treev.heading("2", text ="التليفون")
                treev.heading("3", text ="رقم التسجيل")
                treev.heading("4", text ="الرصيد")
                treev.heading("5", text ="الاسم")
                treev.heading("6", text ="الكود")


                counter = 0
                for sheet in wb.worksheets:
                        counter +=1
                        ws = sheet
                        if (ws.cell(row=1,column=2).value) != None and sheet.title != 'مصروفات اداريه':
                                amount = 0
                                for i in range (4, ws.max_row+1):
                                        try:
                                                amount = int(ws.cell(row=i,column=5).value)
                                        except:
                                                continue   
                                if counter % 2 == 0:             
                                        treev.insert("", 'end', text ='',
                                                        values =(
                                                        '-' if (ws.cell(row=1,column=5).value) == None else str(ws.cell(row=1,column=5).value),
                                                        '-' if (ws.cell(row=1,column=4).value) == None else str(ws.cell(row=1,column=4).value),
                                                        '-' if (ws.cell(row=1,column=3).value) == None else str(ws.cell(row=1,column=3).value),
                                                        '-' if int(amount) == None else int(amount),
                                                        '-' if (ws.cell(row=1,column=1).value) == None else str(ws.cell(row=1,column=1).value),
                                                        '-' if (ws.cell(row=1,column=2).value) == None else str(ws.cell(row=1,column=2).value),
                                                        ), tags = ('even',))
                                else:             
                                        treev.insert("", 'end', text ='',
                                                        values =(
                                                        '-' if (ws.cell(row=1,column=5).value) == None else str(ws.cell(row=1,column=5).value),
                                                        '-' if (ws.cell(row=1,column=4).value) == None else str(ws.cell(row=1,column=4).value),
                                                        '-' if (ws.cell(row=1,column=3).value) == None else str(ws.cell(row=1,column=3).value),
                                                        '-' if int(amount) == None else int(amount),
                                                        '-' if (ws.cell(row=1,column=1).value) == None else str(ws.cell(row=1,column=1).value),
                                                        '-' if (ws.cell(row=1,column=2).value) == None else str(ws.cell(row=1,column=2).value),
                                                        ), tags = ('odd',))
                                                                
                treev.tag_configure('even', background='#e1dddd', font=font22) 
                treev.tag_configure('odd', background='#f5f3f3', font=font22) 


        Button(self, height = 1, width = 15, bg = '#05659E', fg = 'white',
        activebackground='#43516C', font = font20, bd = '8px solid #DBA531', 
                text='ابحث',
                command= search).place(relx=.5, rely=.55,anchor= CENTER)

        Button(self, height = 2, width = 20, bg = '#D85426', fg = 'white',
        activebackground='#D85426', font = 'fantasy 24 bold', bd = '8px solid #DBA531', 
                text='العملاء',
                command= clients).place(relx=.5, rely=.2,anchor= CENTER)        
        
        Button(self, height = 1, width = 10, bg = 'grey', fg = 'white',
        activebackground='#43516C', font = 'fantasy 15 bold', bd = '8px solid #DBA531', 
                text='اغلاق',
                command=self.destroy).place(relx=.5, rely=.8,anchor= CENTER)

# --------------------------------------------------------------------------------------------------------------------------------------------

class Client_Form(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)

        self.geometry(f'{width}x{height}')
        self.title('اضافة عميل')
        self['bg']=baseColor

        Label(self, text='الاسم',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.92, rely=.1,anchor= CENTER, width=int(width*0.138))

        name = tk.StringVar()
        entry1 = ttk.Entry(self, textvariable=name, justify = LEFT, font = font20)
        entry1.place(relx=.77, rely=.17,anchor= CENTER, width=int(width*0.4))

        Label(self, text='.برجاء عدم ترك الخانه فارغه',
        bg = baseColor, fg = labelColor ,font = font10).place(relx=.92, rely=.22,anchor= CENTER)

        Label(self, text='الكود',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.4, rely=.1,anchor= CENTER, width=int(width*0.138))

        code = IntVar()
        entry1 = ttk.Entry(self, textvariable=code, justify = LEFT, font = font20)
        entry1.place(relx=.4, rely=.17,anchor= CENTER, width=int(width*0.138))

        Label(self, text='.برجاء عدم ترك الخانه فارغه',
        bg = baseColor, fg = labelColor ,font = font10).place(relx=.425, rely=.22,anchor= CENTER)


        Label(self, text='رقم التسجيل',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.1, rely=.1,anchor= CENTER, width=int(width*0.138))

        record = tk.IntVar()
        entry1 = ttk.Entry(self, textvariable=record, justify = LEFT, font = font20)
        entry1.place(relx=.1, rely=.17,anchor= CENTER, width=int(width*0.138))


        Label(self, text='رقم التليفون',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.92, rely=.75,anchor= CENTER, width=int(width*0.138))

        phone = tk.StringVar()
        entry1 = ttk.Entry(self, textvariable=phone, justify = LEFT, font = font20)
        entry1.place(relx=.78, rely=.8,anchor= CENTER, width=int(width*0.4))


        Label(self, text='العنوان',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.375, rely=.75,anchor= CENTER, width=int(width*0.138))

        address = tk.StringVar()
        entry1 = ttk.Entry(self, textvariable=address, justify = LEFT, font = font20)
        entry1.place(relx=.25, rely=.8,anchor= CENTER, width=int(width*0.4))


        Label(self, text='الخدمه',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.92, rely=.3,anchor= CENTER, width=int(width*0.138))

        services = []
        ws = wb['الخدمات']
        for row in range(1,ws.max_row+1):
                if ws.cell(column=1, row=row).value != None or ws.cell(column=1, row=row).value != '':
                        services.append(ws.cell(column=1, row=row).value)

        def check_input_services(event):
                value = event.widget.get()
                if value == '':
                        menu2['values'] = services
                else:
                        data = []
                        for item in services:
                                if value.lower() in item.lower():
                                        data.append(item)

                        menu2['values'] = data

        service = tk.StringVar()
        menu2 = ttk.Combobox(self, textvariable = service, font=font20, values = services) 
        app.option_add('*TCombobox*Listbox.font', font20)
        menu2.place(relx=.77, rely=.37,anchor= CENTER, width=int(width*0.4))
        menu2.bind('<KeyRelease>', check_input_services)

        Label(self, text='.برجاء عدم ترك الخانه فارغه',
        bg = baseColor, fg = labelColor ,font = font10).place(relx=.92, rely=.42,anchor= CENTER)
        

        Label(self, text='التكلفه',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.4, rely=.3,anchor= CENTER, width=int(width*0.138))

        cost = IntVar()
        entry1 = ttk.Entry(self, textvariable=cost, justify = LEFT, font = font20)
        entry1.place(relx=.4, rely=.37,anchor= CENTER, width=int(width*0.138))

        Label(self, text='.برجاء عدم ترك الخانه فارغه',
        bg = baseColor, fg = labelColor ,font = font10).place(relx=.425, rely=.42,anchor= CENTER)


        Label(self, text='المدفوع',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.1, rely=.3,anchor= CENTER, width=int(width*0.138))

        amount = IntVar()
        entry1 = ttk.Entry(self, textvariable=amount, justify = LEFT, font = font20)
        entry1.place(relx=.1, rely=.37,anchor= CENTER, width=int(width*0.138))

        Label(self, text='.برجاء عدم ترك الخانه فارغه',
        bg = baseColor, fg = labelColor ,font = font10).place(relx=.125, rely=.42,anchor= CENTER)
        
        
        Label(self, text='ملاحظات',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.92, rely=.5,anchor= CENTER, width=int(width*0.138))

        comment = Text(self, font = font20, height=3)
        comment.place(relx=.78, rely=.62,anchor= CENTER, width=int(width*0.4),)


        Label(self, text='التاريخ',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.375, rely=.5,anchor= CENTER, width=int(width*0.138))

        c_date = DateEntry(self,width=int(width*0.02),bg="darkblue",fg="white", font = font18,
        year=datetime.now().year,month=datetime.now().month,day=datetime.now().day
        ,locale='en_US', date_pattern='dd/MM/yyyy')
        c_date.place(relx=.25, rely=.57,anchor= CENTER, width=int(width*0.4))
        
        
        
        entries = []
        for sheet in wb.worksheets:
                if sheet.cell(row=1,column=1).value != None or sheet.title != 'مصروفات اداريه' or sheet.title != 'الخدمات':
                        entries.append(str(sheet.cell(row=1,column=1).value)) 

        codeentries = []
        for sheet in wb.worksheets:
                if sheet.cell(row=1,column=2).value != None :
                        codeentries.append(sheet.cell(row=1,column=2).value)                

        def insertClient():
                # get() validation
                try:
                        name.get() and code.get() and record.get() and service.get() and cost.get() and amount.get() and phone.get() and address.get() and comment.get()
                except:
                        messagebox.showerror('Invalid!','من فضلك قم بادخال الخانات بطريقه صحيحه ') 
                        print('Invalid Inputs')
                        return False

                # not None Validation
                if name.get() == None or code.get() == None or service.get() == None or cost.get() == None or amount.get() == None :
                        print('None Inputs')        
                        return False

                # not "" validation        
                if name.get() == "" or service.get() == "" :
                        messagebox.showerror('Invalid!','من فضلك نأكد من ادخال الاسم و الخدمه ') 
                        print('No Name or Service')
                        return False

                # phone validation
                if not phone.get().isnumeric:
                        # self.destroy()
                        messagebox.showerror('Invalid!','من فضلك نأكد من ادخال التليفون بطريقه صحيحه ') 
                        print('invalid Phone')  
                        return False

                # Case validations
                if name.get() in entries :
                        messagebox.showerror('Already Exists!','الاسم موجود بالفعل ') 
                        print('Name Exists')
                        return False

                if code.get() in codeentries:
                        messagebox.showerror('Already Exists!','الكود موجود بالفعل ') 
                        print('Code Exists')
                        return False
                
                if len(name.get()) > 26:
                        messagebox.showerror('Error!','من فضلك استخدم اسم لا تزيد حروفه عن 23 حرف') 
                        print('Name is long')
                        return False
                
                else:  
                        # Create Client sheet      
                        ws = wb.create_sheet(f'{name.get()} ({code.get()})')
                        ws.title = f'{name.get()} ({code.get()})'
                        ws.append([name.get(), code.get(),str(record.get()),phone.get(),address.get()])
                        ws.append([])
                        ws.append(['التاريخ', 'الخدمه','التكلفه','المدفوع','الرصيد', 'الملاحظات',])
                        ws.append([c_date.get_date().strftime("%d/%m/%Y"),
                        service.get(),
                        cost.get(),
                        amount.get(), 
                        (cost.get()) - amount.get() ,
                        comment.get("1.0",END),
                        c_date.get_date().year,c_date.get_date().month, c_date.get_date().day])
                        wb.save(file)

                        # Create Clients Expenses sheet
                        ws0 = wb0.create_sheet(f'{name.get()} ({code.get()})')
                        ws0.title = f'{name.get()} ({code.get()})'
                        ws0.append([name.get(), code.get(),str(record.get()),phone.get(),address.get()])
                        ws0.append([])
                        ws0.append(['التاريخ', 'المبلغ','المصروف'])
                        wb0.save(file0)
                        self.destroy()
                        messagebox.showinfo('Done','تم الحفظ بنجاح ')

        
        Button(self, height = 1, width = 10, bg = '#05659E', fg = 'white',
        activebackground='#43516C', font = font20, bd = '4px solid #DBA531', 
                text='حفظ',
                command=insertClient).place(relx=.5, rely=.9,anchor= CENTER)
        
        
        Button(self, height = 1, width = 8, bg = 'grey', fg = 'white',
        activebackground='#43516C', font = 'fantasy 15 bold', bd = '4px solid #DBA531', 
                text='اغلاق',
                command=self.destroy).place(relx=.1, rely=.95,anchor= CENTER)

# -------------------------------------------------------------------------------------------------------

class Service_Form(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)

        self.geometry(f'{width}x{height}')
        self.title('اضافة خدمه')
        self['bg']=baseColor

        entries = []
        for sheet in wb.worksheets:
                if sheet.cell(row=1,column=1).value != None or sheet.title != 'الخدمات':
                        entries.append(f'{str(sheet.cell(row=1,column=1).value)} ({str(sheet.cell(row=1,column=2).value)})')
                        
        Label(self, text='الاسم',
        bg = baseColor, fg = labelColor ,font = font20 ).place(relx=.8, rely=.1,anchor= CENTER, width=int(width*0.138))


        def check_input_names(event):
                value = event.widget.get()
                if value == '':
                        menu['values'] = entries
                else:
                        data = []
                        for item in entries:
                                if value.lower() in item.lower():
                                        data.append(item)

                        menu['values'] = data

        name = StringVar()
        menu = ttk.Combobox(self, textvariable= name, font=font20,values = entries )
        app.option_add('*TCombobox*Listbox.font', font20)
        menu.place(relx=.55, rely=.1,anchor= CENTER, width=int(width*0.417))
        menu['values'] = entries
        menu.bind('<KeyRelease>', check_input_names)

        Label(self, text='.برجاء عدم ترك الخانه فارغه',
        bg = baseColor, fg = labelColor ,font = font10).place(relx=.71, rely=.15,anchor= CENTER)


        Label(self, text='الخدمه',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.8, rely=.22,anchor= CENTER, width=int(width*0.138))

        services = []
        ws = wb['الخدمات']
        for row in range(1, ws.max_row+1):
                if ws.cell(column=1, row=row).value != None or ws.cell(column=1, row=row).value != '':
                        services.append(ws.cell(column=1, row=row).value)

        
        def check_input_services(event):
                value = event.widget.get()
                if value == '':
                        menu2['values'] = services
                else:
                        data = []
                        for item in services:
                                if value.lower() in item.lower():
                                        data.append(item)

                        menu2['values'] = data
                        
        service = tk.StringVar()
        menu2 = ttk.Combobox(self, textvariable = service, font=font20, values = services)
        app.option_add('*TCombobox*Listbox.font', font20)
        menu2.place(relx=.55, rely=.22,anchor= CENTER, width=int(width*0.417))
        menu2.bind('<KeyRelease>', check_input_services)

        Label(self, text='.برجاء عدم ترك الخانه فارغه',
        bg = baseColor, fg = labelColor ,font = font10).place(relx=.71, rely=.27,anchor= CENTER)
        

        Label(self, text='التكلفه',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.8, rely=.34,anchor= CENTER, width=int(width*0.138))

        cost = IntVar()
        entry1 = ttk.Entry(self, textvariable=cost, justify = LEFT, font = font20)
        entry1.place(relx=.55, rely=.34,anchor= CENTER, width=int(width*0.417), height=50)

        Label(self, text='.برجاء عدم ترك الخانه فارغه',
        bg = baseColor, fg = labelColor ,font = font10).place(relx=.71, rely=.392,anchor= CENTER)


        Label(self, text='المدفوع',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.8, rely=.58,anchor= CENTER, width=int(width*0.138))

        amount = IntVar()
        entry1 = ttk.Entry(self, textvariable=amount, justify = LEFT, font = font20)
        entry1.place(relx=.55, rely=.58,anchor= CENTER, width=int(width*0.417), height=int(height*0.06))

        Label(self, text='.برجاء عدم ترك الخانه فارغه',
        bg = baseColor, fg = labelColor ,font = font10).place(relx=.715, rely=.63,anchor= CENTER)
        

        Label(self, text='ملاحظات',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.8, rely=.7,anchor= CENTER, width=int(width*0.138))

        comment =  Text(self, font = font20, height=3)
        comment.place(relx=.55, rely=.73,anchor= CENTER, width=int(width*0.417))


        Label(self, text='التاريخ',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.8, rely=.46,anchor= CENTER, width=int(width*0.138))

        c_date = DateEntry(self,width=30,bg="darkblue",fg="white", font = font18,
        year=datetime.now().year,month=datetime.now().month,day=datetime.now().day
        ,locale='en_US', date_pattern='dd/MM/yyyy')
        c_date.place(relx=.675, rely=.46,anchor= CENTER, width=int(width*0.138), height=int(height*0.06))
        
        
        def addService():
                try:
                        amount.get(), cost.get(), service.get(), name.get() 
                except:
                        messagebox.showerror('Invalid!','من فضلك قم بادخال الخانات بطريقه صحيحه ')
                        print('Invalid inputs')
                        return False

                if cost.get() == '' or name.get() == '' or service.get() == '':
                        messagebox.showerror('Invalid!','من فضلك قم بادخال الخانات المطلوبه ')
                        print('Empty Inputs')
                        return False
                try:
                        cost.get() 
                except:
                        messagebox.showerror('Invalid!','من فضلك قم بادخال خانة التكلفه بطريقه صحيحه ')
                        print('Invalid Code')
                        return False

                try:
                        service.get() 
                except:
                        messagebox.showerror('Invalid!','من فضلك قم بادخال خانة الخدمه بطريقه صحيحه ')  
                        print('Invalid Service')
                        return False
                        
                try:
                        ws = wb[name.get()]
                except:
                        messagebox.showerror('Not Exists!','الاسم غير موجود')
                        print('')
                        return False         

                inserted_date = date(c_date.get_date().year,c_date.get_date().month, c_date.get_date().day)
                exists_dates = []
                # check if transaction needs to be sorted by comparing date
                for i in range (4, ws.max_row+1):
                        try:
                                exists_dates.append (date(
                                ws.cell(column=7, row=i).value,
                                ws.cell(column=8, row=i).value,
                                ws.cell(column=9, row=i).value,))
                        except:
                                continue
                # if transaction needs to be sorted
                flag = False
                for cell_date in exists_dates:
                        if inserted_date < cell_date:
                                # add all transactions to rows_list
                                rows_list = []
                                for i in range (4, ws.max_row+1):
                                        try:
                                                cur_row_date = date(
                                                        ws.cell(column=7, row=i).value,
                                                        ws.cell(column=8, row=i).value,
                                                        ws.cell(column=9, row=i).value,)
                                        except:
                                                continue        
                                        cur_row = [    
                                        ws.cell(column=1, row=i).value,
                                        ws.cell(column=2, row=i).value,
                                        ws.cell(column=3, row=i).value,
                                        ws.cell(column=4, row=i).value,
                                        ws.cell(column=5, row=i).value,
                                        ws.cell(column=6, row=i).value,
                                        ws.cell(column=7, row=i).value,
                                        ws.cell(column=8, row=i).value,
                                        ws.cell(column=9, row=i).value,
                                        cur_row_date]

                                        rows_list.append(cur_row)
                                # add the new transaction
                                rows_list.append([
                                c_date.get_date().strftime("%d/%m/%Y"),
                                service.get(),
                                cost.get(),
                                amount.get(),
                                (cost.get() - amount.get()),
                                comment.get("1.0",END),
                                c_date.get_date().year,c_date.get_date().month, c_date.get_date().day,
                                inserted_date,])  

                                # sort rows_list by date
                                rows_list.sort(key=lambda x : x[9])
                                # delete rows 
                                ws.delete_rows(4, ws.max_row)
                                #add sorted rows
                                
                                ws.append([rows_list[0][0],rows_list[0][1],rows_list[0][2],rows_list[0][3],int(rows_list[0][2])-int(rows_list[0][3]),rows_list[0][5],rows_list[0][6],rows_list[0][7],rows_list[0][8],rows_list[0][9]])
                                
                                rows_list.pop(0)
                                for row in rows_list:
                                        try:
                                                ws.append([row[0],row[1],row[2],row[3],row[2]-row[3]+int(ws.cell(column=5,row=ws.max_row).value),row[5],row[6],row[7],row[8],row[9],])
                                        except:
                                                continue
                                wb.save(file)
                                flag = True
                                # self.destroy()
                                messagebox.showinfo('Done','تم الحفظ بنجاح ')
                                break 

                # if transaction doesn't need to be sorted
                if flag == False:
                        finalAmount = ws.cell(column=5,row=int(ws.max_row)).value
                        if type(finalAmount) == str or finalAmount == 0 or finalAmount == None:
                                finalAmount = 0           
                        ws.append([c_date.get_date().strftime("%d/%m/%Y"),
                        service.get(),
                        cost.get(),
                        amount.get(),
                        ((int(finalAmount) + (cost.get()) - amount.get())),
                        comment.get("1.0",END),
                        c_date.get_date().year,c_date.get_date().month, c_date.get_date().day,
                        inserted_date,])

                        wb.save(file)
                        # self.destroy()
                        messagebox.showinfo('Done','تم الحفظ بنجاح ')

        
        Button(self, height = 1, width = 10, bg = '#05659E', fg = 'white',
        activebackground='#43516C', font = font20, bd = '4px solid #DBA531', 
                text='حفظ',
                command=addService).place(relx=.5, rely=.9,anchor= CENTER)
        
        Button(self, height = 1, width = 8, bg = 'grey', fg = 'white',
        activebackground='#43516C', font = 'fantasy 15 bold', bd = '4px solid #DBA531', 
                text='اغلاق',
                command=self.destroy).place(relx=.1, rely=.95,anchor= CENTER)

# -----------------------------------------------------------------------------------------------------------------------------

class Delete(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)

        self.geometry(f'{width}x{height}')
        self.title('حذف عميل')
        self['bg']=baseColor

        entries = []
        for sheet in wb.worksheets:
                if sheet.cell(row=1,column=1).value != None or sheet.title != 'الخدمات':
                        entries.append(f'{str(sheet.cell(row=1,column=1).value)} ({str(sheet.cell(row=1,column=2).value)})') 


        def check_input(event):
                value = event.widget.get()
                if value == '':
                        menu['values'] = entries
                else:
                        data = []
                        for item in entries:
                                if value.lower() in item.lower():
                                        data.append(item)

                        menu['values'] = data


        Label(self, text='اسم العميل',
        bg = baseColor, fg = labelColor ,font = font20).place(relx=.72, rely=.3,anchor= CENTER)

        name = StringVar()
        menu = ttk.Combobox(self, textvariable= name, font=font20,values = entries )
        app.option_add('*TCombobox*Listbox.font', font20)
        menu.place(relx=.5, rely=.35,anchor= CENTER, width=int(width*0.55))
        menu.bind('<KeyRelease>', check_input)

        Label(self, text='.برجاء عدم ترك الخانه فارغه',
        bg = baseColor, fg = labelColor ,font = font10).place(relx=.73, rely=.4,anchor= CENTER)

        def delete():
                response = messagebox.askquestion('Deleted!', f'{name.get()} هل تريد ان تحذف كشف حساب العميل')
                if response == 'yes' :
                        try:
                                del wb[name.get()]
                                del wb0[name.get()]
                        except:
                                messagebox.showerror('Not Exists!','الاسم غير موجود') 
                                print('Invalid Name')
                                return False        
                        self.destroy()
                        messagebox.showinfo('Done!','تم الحذف بنجاح')
                        wb.save(file)
                        wb0.save(file0)
                else :
                        messagebox.showinfo('Fail!', 'لم يتم الحذف')
                        return False       


        Button(self, height = 1, width = 15, bg = 'red', fg = 'white',
        activebackground='#43516C', font = font20, bd = '8px solid #DBA531', 
                text='حذف',
                command=delete).place(relx=.5, rely=.5,anchor= CENTER)
        
        Button(self, height = 1, width = 10, bg = 'grey', fg = 'white',
        activebackground='#43516C', font = 'fantasy 15 bold', bd = '8px solid #DBA531', 
                text='اغلاق',
                command=self.destroy).place(relx=.5, rely=.8,anchor= CENTER)


class Main(tk.Tk):
    def __init__(self):
        super().__init__()

        self.geometry(f'{width}x{height}')
        self.title('Main Window')
        self['bg']=baseColor

        Label(self, text='Mr.Wagdy Latif For Accounting Services',
        bg = '#D85426', fg = 'white', font = 'fantasy 30 bold', borderwidth=20, relief="ridge", padx=20, pady=20).place(relx=.5, rely=.1,anchor= CENTER)
                
        Button(self, height = 1, width = 10, bg = 'green', fg = 'white',
        activebackground='green', font = 'fantasy 25 bold', bd = '8px solid #DBA531', 
                text='عملاء',
                command=self.open_insert).place(relx=.8, rely=.35,anchor= CENTER)

        Button(self, height = 1, width = 10, bg = 'green', fg = 'white',
        activebackground='green', font = 'fantasy 25 bold', bd = '8px solid #DBA531', 
                text='خدمات',
                command=self.open_add).place(relx=.8, rely=.5,anchor= CENTER) 

        Button(self, height = 1, width = 10, bg = 'green', fg = 'white',
        activebackground='green', font = 'fantasy 25 bold', bd = '8px solid #DBA531', 
                text='مدفوعات',
                command=self.open_payment).place(relx=.8, rely=.65,anchor= CENTER)

        Button(self, height = 1, width = 10, bg = 'green', fg = 'white',
        activebackground='green', font = 'fantasy 25 bold', bd = '8px solid #DBA531', 
                text='مصروفات',
                command=self.open_exp_form).place(relx=.8, rely=.8,anchor= CENTER) 

        Button(self, height = 2, width = 12, bg = '#05659E', fg = 'white',
        activebackground='#43516C', font = 'fantasy 25 bold', bd = '8px solid #DBA531', 
                text='العملاء',
                command=self.open_search).place(relx=.5, rely=.35,anchor= CENTER)

        Button(self, height = 2, width = 12, bg = '#05659E', fg = 'white',
        activebackground='#43516C', font = 'fantasy 25 bold', bd = '8px solid #DBA531', 
                text='الإيرادات',
                command=self.open_revenue).place(relx=.5, rely=.6,anchor= CENTER)       

        Button(self, height = 2, width = 12, bg = '#05659E', fg = 'white',
        activebackground='#43516C', font = 'fantasy 25 bold', bd = '8px solid #DBA531', 
                text= 'المصروفات',
                command=self.open_exp).place(relx=.5, rely=.85,anchor= CENTER)        

        Button(self, height = 1, width = 12, bg = 'red', fg = 'white',
        activebackground='#43516C', font = 'fantasy 25 bold', bd = '8px solid #DBA531', 
                text='حذف',
                command=self.open_delete).place(relx=.2, rely=.85,anchor= CENTER)
      
        Button(self, height = 1, width = 12, bg = '#F1B93F', fg = labelColor,
        activebackground='#43516C', font = 'fantasy 25 bold', bd = '8px solid #DBA531', 
                text='تعديل ملفات العملاء',
                command=self.open_edit_client).place(relx=.2, rely=.6,anchor= CENTER) 

        Button(self, height = 1, width = 12, bg = '#F1B93F', fg = labelColor,
        activebackground='#43516C', font = 'fantasy 25 bold', bd = '8px solid #DBA531', 
                text='تعديل قائمة الخدمات',
                command=self.open_new_service).place(relx=.2, rely=.35,anchor= CENTER)                


    def open_new_service(self):
        window = New_Service(self)
        window.grab_set() 

    def open_edit_client(self):
        window = Edit_Client(self)
        window.grab_set() 

    def open_exp(self):
                window = Expenses_View(self)
                window.grab_set() 

    def open_revenue(self):
                window = Revenues_View(self)
                window.grab_set()  

    def open_payment(self):
                window = Payment_Form(self)
                window.grab_set()            

    def open_search(self):
                window = Search(self)
                window.grab_set()

    def open_insert(self):
                window = Client_Form(self)
                window.grab_set()

    def open_delete(self):
                window = Delete(self)
                window.grab_set()  

    def open_add(self):
                window = Service_Form(self)
                window.grab_set()
                
    def open_exp_form(self):
                window = Expenses_Form(self)
                window.grab_set()           

if __name__ == "__main__":
        app = Main()
        app.mainloop()
